#!/usr/bin/env python
# -*- coding: utf8 -*-
# ============================================================================
#  Copyright (c) nexB Inc. http://www.nexb.com/ - All rights reserved.
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#      http://www.apache.org/licenses/LICENSE-2.0
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
# ============================================================================

from __future__ import absolute_import
from __future__ import print_function
from __future__ import unicode_literals

import codecs
from collections import OrderedDict
import io
import json
import ntpath
import openpyxl
import os
import posixpath
import re
import shutil
import string
import sys
import yaml


from attributecode import CRITICAL
from attributecode import ERROR
from attributecode import WARNING
from attributecode import Error
from attributecode import model

from itertools import zip_longest  # NOQA

import csv  # NOQA

on_windows = 'win32' in sys.platform

file_fields = ['notice_file', 'license_file']

def to_posix(path):
    """
    Return a path using the posix path separator given a path that may contain
    posix or windows separators, converting "\\" to "/". NB: this path will
    still be valid in the windows explorer (except for a UNC or share name). It
    will be a valid path everywhere in Python. It will not be valid for windows
    command line operations.
    """
    return path.replace(ntpath.sep, posixpath.sep)


UNC_PREFIX = u'\\\\?\\'
UNC_PREFIX_POSIX = to_posix(UNC_PREFIX)
UNC_PREFIXES = (UNC_PREFIX_POSIX, UNC_PREFIX,)

valid_file_chars = string.digits + string.ascii_letters + '_-.+()~[]{}|@' + ' '


"""
Return True if a string s  name is safe to use as an attribute name.
"""
is_valid_name = re.compile(r'^[A-Za-z_][A-Za-z0-9_]*$').match


def load_csv(location, configuration=None):
    """
    Read CSV at `location`, return a list of ordered dictionaries, one
    for each row.
    """
    results = []
    mapping_dict = {}
    if configuration:
        with open(configuration) as file:
            mapping_dict = yaml.load(file)
    # FIXME: why ignore encoding errors here?
    with codecs.open(location, mode='rb', encoding='utf-8-sig',
                     errors='ignore') as csvfile:
        for row in csv.DictReader(csvfile):
            # convert all the column keys to lower case
            updated_row = OrderedDict()
            for key, value in row.items():
                if key in mapping_dict:
                    key = mapping_dict[key]
                updated_row[key.lower()] = value
            results.append(updated_row)
    return results


def load_excel(location, configuration=None):
    """
    Read Excel at `location`, return a list of ordered dictionaries, one
    for each row.
    """
    results = []
    errors = []
    sheet_obj = openpyxl.load_workbook(location).active
    max_col = sheet_obj.max_column

    index = 1
    col_keys = []
    mapping_dict = {}
    if configuration:
        with open(configuration) as file:
            mapping_dict = yaml.load(file)
    while index <= max_col:
        value = sheet_obj.cell(row=1, column=index).value
        if value in col_keys:
            msg = 'Duplicated column name, ' + str(value) + ', detected.' 
            errors.append(Error(CRITICAL, msg))
            return errors, results
        if value in mapping_dict:
            value = mapping_dict[value]
        col_keys.append(value)
        index = index + 1

    for row in sheet_obj.iter_rows(min_row=2, values_only=True):
        row_dict = OrderedDict()
        index = 0
        while index < max_col:
            value = row[index]
            if value:
                row_dict[col_keys[index]] = value
            else:
                row_dict[col_keys[index]] = ''
            index = index + 1
        results.append(row_dict)
    return errors, results


def load_scancode_json(location, configuration=None):
    """
    Read the scancode JSON file at `location` and return a list of dictionaries.
    """
    mapping_dict = {}
    updated_results = []
    if configuration:
        with open(configuration) as file:
            mapping_dict = yaml.load(file)
    with open(location) as json_file:
        results = json.load(json_file)
    results = results['files']
    if mapping_dict:
        for item in results:
            updated_item = {}
            for key in item:
                if key in mapping_dict:
                    updated_item[mapping_dict[key]] = item[key]
                else:
                    updated_item[key] = item[key]
            updated_results.append(updated_item)
    else:
        updated_results = results
    return updated_results


def load_json(location):
    """
    Read JSON file at `location` and return a list of ordered dicts, one for
    each entry.
    """
    with open(location) as json_file:
        results = json.load(json_file)
    if isinstance(results, list):
        results = sorted(results)
    else:
        results = [results]
    return results


# FIXME: rename to is_online: BUT do we really need this at all????
def have_network_connection():
    """
    Return True if an HTTP connection to some public web site is possible.
    """
    import socket
    import http.client as httplib  # NOQA

    http_connection = httplib.HTTPConnection('dejacode.org', timeout=10)  # NOQA
    try:
        http_connection.connect()
    except socket.error:
        return False
    else:
        return True

def add_unc(location):
    """
    Convert a `location` to an absolute Window UNC path to support long paths on
    Windows. Return the location unchanged if not on Windows. See
    https://msdn.microsoft.com/en-us/library/aa365247.aspx
    """
    if on_windows and not location.startswith(UNC_PREFIX):
        if location.startswith(UNC_PREFIX_POSIX):
            return UNC_PREFIX + os.path.abspath(location.strip(UNC_PREFIX_POSIX))
        return UNC_PREFIX + os.path.abspath(location)
    return location

def unique(sequence):
    """
    Return a list of unique items found in sequence. Preserve the original
    sequence order.
    For example:
    >>> unique([1, 5, 3, 5])
    [1, 5, 3]
    """
    deduped = []
    for item in sequence:
        if item not in deduped:
            deduped.append(item)
    return deduped


def filter_errors(errors, minimum_severity=WARNING):
    """
    Return a list of unique `errors` Error object filtering errors that have a
    severity below `minimum_severity`.
    """
    return unique([e for e in errors if e.severity >= minimum_severity])


def create_dir(location):
    """
    Create directory or directory tree at location, ensuring it is readable
    and writeable.
    """
    import stat
    if not os.path.exists(location):
        os.makedirs(location)
        os.chmod(location, stat.S_IRWXU | stat.S_IRWXG
                 | stat.S_IROTH | stat.S_IXOTH)


def get_temp_dir(sub_dir_path=None):
    """
    Create a unique new temporary directory location. Create directories
    identified by sub_dir_path if provided in this temporary directory.
    Return the location for this unique directory joined with the
    sub_dir_path if any.
    """
    new_temp_dir = build_temp_dir()

    if sub_dir_path:
        # create a sub directory hierarchy if requested
        new_temp_dir = os.path.join(new_temp_dir, sub_dir_path)
        create_dir(new_temp_dir)
    return new_temp_dir


def build_temp_dir(prefix='attributecode-'):
    """
    Create and return a new unique empty directory created in base_dir.
    """
    import tempfile
    location = tempfile.mkdtemp(prefix=prefix)
    create_dir(location)
    return location

def get_file_text(file_name, reference):
    """
    Return the file content from the license_file/notice_file field from the
    given reference directory.
    """
    error = ''
    text = ''
    file_path = os.path.join(reference, file_name)
    if not os.path.exists(file_path):
        msg = "The file " + file_path + " does not exist"
        error = Error(CRITICAL, msg)
    else:
        with codecs.open(file_path, 'rb', encoding='utf-8-sig', errors='replace') as txt:
        #with io.open(file_path, encoding='utf-8') as txt:
            text = txt.read()
    return error, text

def check_duplicated_columns(location):
    """
    Return a list of errors for duplicated column names in a CSV file
    at location.
    """
    location = add_unc(location)
    with codecs.open(location, 'rb', encoding='utf-8-sig', errors='replace') as csvfile:
        reader = csv.reader(csvfile)
        columns = next(reader)
        columns = [col for col in columns]

    seen = set()
    dupes = OrderedDict()
    for col in columns:
        c = col.lower()
        if c in seen:
            if c in dupes:
                dupes[c].append(col)
            else:
                dupes[c] = [col]
        seen.add(c.lower())

    errors = []
    if dupes:
        dup_msg = []
        for name, names in dupes.items():
            names = u', '.join(names)
            msg = '%(name)s with %(names)s' % locals()
            dup_msg.append(msg)
        dup_msg = u', '.join(dup_msg)
        msg = ('Duplicated column name(s): %(dup_msg)s\n' % locals() +
               'Please correct the input and re-run.')
        errors.append(Error(ERROR, msg))
    return unique(errors)

def check_newline_in_file_field(component):
    """
    Return a list of errors for newline characters detected in *_file fields.
    """
    errors = []
    for k in component.keys():
        if k in file_fields:
            try:
                if '\n' in component[k]:
                    msg = ("New line character detected in '%s' for '%s' which is not supported."
                            "\nPlease use ',' to declare multiple files.") % (k, component['about_resource'])
                    errors.append(Error(CRITICAL, msg))
            except:
                pass
    return errors

def load_inventory(location, configuration=None, scancode=False, reference_dir=None):
    """
    Load the inventory file at `location` 

    Optionally use `reference_dir` as the directory location of extra reference
    license and notice files to reuse.
    """
    errors = []
    abouts = []
    if scancode:
        inventory = load_scancode_json(location, configuration)
    else:
        if location.endswith('.csv'):
            dup_cols_err = check_duplicated_columns(location)
            if dup_cols_err:
                errors.extend(dup_cols_err)
                return errors, abouts
            inventory = load_csv(location, configuration)
        elif location.endswith('.xlsx'):
            dup_cols_err, inventory = load_excel(location, configuration)
            if dup_cols_err:
                errors.extend(dup_cols_err)
                return errors, abouts
        else:
            inventory = load_json(location)

    errors = []
    for component in inventory:
        newline_in_file_err = check_newline_in_file_field(component)
        for err in newline_in_file_err:
            errors.append(err)

    if errors:
        return errors, abouts

    for component in inventory:
        about = model.About()
        ld_errors = about.load_dict(
            component,
            scancode=scancode,
            reference_dir=reference_dir,
        )
        for e in ld_errors:
            if not e in errors:
                errors.extend(ld_errors)
        abouts.append(about)

    return unique(errors), abouts

def convert_object_to_dict(about):
    """
    Convert the list of field object
        [Field(name='name', value=''), Field(name='version', value='')]
    to a dictionary
    """
    about_dict = {}
    # Convert all the supported fields into a dictionary
    fields_dict = getattr(about, 'fields')
    custom_fields_dict = getattr(about, 'custom_fields')
    supported_dict = {**fields_dict, **custom_fields_dict}
    for field in supported_dict:
        key = supported_dict[field].name
        value = supported_dict[field].value
        about_dict[key] = value
    return about_dict

def number_of_component_generated_from_default_template(location):
    """
    Return number of component generated from the default template.
    """
    lines = []
    with open(location) as f:
        lines = f.readlines()
    count = 0
    for line in lines:
        if '<h3 class="component-name">' in line:
            if line.replace('<h3 class="component-name">', '').strip():
                count += 1
    return count
